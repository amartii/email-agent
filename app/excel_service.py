"""
excel_service.py
Handles reading contacts from Excel and writing status updates back.
"""
import os
import shutil
from datetime import datetime
from typing import Optional

import openpyxl
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename
from flask import current_app


STATUS_COLUMN = "Estado"
SENT_AT_COLUMN = "Fecha Envío"
REPLIED_AT_COLUMN = "Fecha Respuesta"
FOLLOWUP_COLUMN = "Follow-up Enviado"

# Columns automatically managed by the app
MANAGED_COLUMNS = [STATUS_COLUMN, SENT_AT_COLUMN, REPLIED_AT_COLUMN, FOLLOWUP_COLUMN]

STATUS_LABELS = {
    "pending": "Pendiente",
    "sent": "Enviado",
    "replied": "Respondido",
    "followup_sent": "Follow-up enviado",
    "bounced": "Rebotado",
}


def save_upload(file: FileStorage) -> str:
    """Save uploaded Excel to the uploads folder. Returns the saved file path."""
    filename = secure_filename(file.filename)
    if not filename.endswith((".xlsx", ".xls")):
        raise ValueError("El archivo debe ser .xlsx o .xls")
    upload_folder = current_app.config["UPLOAD_FOLDER"]
    os.makedirs(upload_folder, exist_ok=True)
    dest = os.path.join(upload_folder, filename)
    file.save(dest)
    return dest


def read_columns(file_path: str) -> list[str]:
    """Return the list of column headers in the Excel file (excludes managed columns)."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        val = cell.value
        if val and str(val).strip() and str(val).strip() not in MANAGED_COLUMNS:
            headers.append(str(val).strip())
    wb.close()
    return headers


def read_contacts(file_path: str, name_col: str, email_col: str) -> list[dict]:
    """
    Read all rows from the Excel.
    Returns list of dicts with keys: 'name', 'email', 'custom_fields' (all other columns).
    Skips rows without a valid email.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    contacts = []

    for row in rows[1:]:
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
        email = row_dict.get(email_col)
        name = row_dict.get(name_col, "")
        if not email or "@" not in str(email):
            continue
        custom = {
            k: v for k, v in row_dict.items()
            if k not in (name_col, email_col) and k not in MANAGED_COLUMNS
        }
        contacts.append({
            "name": str(name).strip() if name else "",
            "email": str(email).strip(),
            "custom_fields": custom,
        })

    wb.close()
    return contacts


def _ensure_managed_columns(ws, headers: list) -> dict:
    """
    Add managed columns (Estado, Fecha Envío, etc.) to the sheet if missing.
    Returns a dict mapping column name → column index (1-based).
    """
    col_map = {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}
    next_col = ws.max_column + 1

    for col_name in MANAGED_COLUMNS:
        if col_name not in col_map:
            ws.cell(row=1, column=next_col, value=col_name)
            col_map[col_name] = next_col
            next_col += 1

    return col_map


def _find_email_col_index(ws, email_col: str) -> Optional[int]:
    """Find the column index for the email field."""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip() == email_col:
            return cell.column
    return None


def update_contact_status(
    file_path: str,
    email: str,
    status: str,
    email_col: str = "Email",
    sent_at: Optional[datetime] = None,
    replied_at: Optional[datetime] = None,
    followup_sent_at: Optional[datetime] = None,
):
    """
    Update status columns for a contact row identified by email.
    Adds managed columns if they don't exist yet. Saves the file in place.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    col_map = _ensure_managed_columns(ws, [])
    email_col_idx = _find_email_col_index(ws, email_col)

    if email_col_idx is None:
        wb.close()
        return

    for row in ws.iter_rows(min_row=2):
        cell_email = row[email_col_idx - 1]
        if cell_email.value and str(cell_email.value).strip().lower() == email.lower():
            row_num = cell_email.row
            ws.cell(row=row_num, column=col_map[STATUS_COLUMN], value=STATUS_LABELS.get(status, status))
            if sent_at:
                ws.cell(row=row_num, column=col_map[SENT_AT_COLUMN], value=sent_at.strftime("%d/%m/%Y %H:%M"))
            if replied_at:
                ws.cell(row=row_num, column=col_map[REPLIED_AT_COLUMN], value=replied_at.strftime("%d/%m/%Y %H:%M"))
            if followup_sent_at:
                ws.cell(row=row_num, column=col_map[FOLLOWUP_COLUMN], value=followup_sent_at.strftime("%d/%m/%Y %H:%M"))
            break

    wb.save(file_path)
    wb.close()
